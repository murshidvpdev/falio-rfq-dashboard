"""
Document field extraction service.
Supports:  .pdf (pdfplumber)  .xlsx/.xls (openpyxl)  .csv  .docx  any (fallback)
"""
import io, csv, re, json, random
from datetime import datetime
from pathlib import Path
from typing import Optional


# ── Column name → field mapping (Excel / CSV) ─────────────────────────────────

_COL_MAP = {
    "description": ["desc", "description", "item desc", "material", "name", "item name", "particulars", "details", "commodity"],
    "quantity":    ["qty", "quantity", "unit qty", "no of units"],
    "unit":        ["unit", "uom", "u/m", "unit of measure"],
    "unit_price":  ["unit price", "price", "rate", "unit rate", "unit cost"],
    "total_price": ["total", "total price", "total amount", "value", "line total", "amount", "extended"],
    "part_number": ["part no", "part number", "part#", "material no", "material code"],
    "item_no":     ["item no", "item #", "s.no", "sno", "sr no", "#", "sl no"],
    "notes":       ["notes", "remarks", "remark", "comment", "alternative"],
}

# ── PDF field → model field ────────────────────────────────────────────────────

_PDF_FIELD_MAP = {
    # label patterns (lowercased)   →  output key
    "owner":              "owner",
    "event type":         "event_type",
    "commodity":          "commodity",
    "currency":           "currency_name",
    "due date":           "due_date",
    "closing date":       "due_date",
    "publish time":       "issue_date",
    "publish date":       "issue_date",
    "received date":      "issue_date",
    "start date":         "issue_date",
    "address":            "address",
    "location":           "address",
    "response id":        "response_id",
    "po id":              "po_id",
    "po number":          "po_number",
    "purchase order":     "po_number",
    "agreement no":       "agreement_no",
    "agreement number":   "agreement_no",
    "contact":            "contact",
    "phone":              "phone",
    "email":              "email",
    "bid date":           "due_date",
    "submission date":    "due_date",
}


def _safe_float(v) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _match_col(header: str) -> Optional[str]:
    h = header.lower().strip()
    for field, variants in _COL_MAP.items():
        if any(v in h for v in variants):
            return field
    return None


def _match_pdf_field(label: str) -> Optional[str]:
    l = label.lower().strip().rstrip(":")
    for pattern, field in _PDF_FIELD_MAP.items():
        if pattern in l:
            return field
    return None


def _extract_rfq_number(text: str, filename: str) -> Optional[str]:
    """Pull SEC-style RFQ number like C001776715 from text or filename."""
    # SEC format: C followed by 9+ digits
    match = re.search(r'\bC\d{7,}\b', filename + " " + text)
    if match:
        return match.group().upper()
    # Generic RFQ/REQ number
    match = re.search(r'(?:RFQ|REQ|RQ|PO)[- #]*([A-Z0-9\-]{5,})', text, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    # Numeric from filename
    match = re.search(r'\d{6,}', filename)
    if match:
        return f"RFQ-{match.group()}"
    return None


def _extract_case_id(filename: str, text: str) -> Optional[str]:
    """Extract case/event ID like 54010 from --54010 pattern."""
    match = re.search(r'--(\d+)', filename)
    if match:
        return match.group(1)
    match = re.search(r'(?:Case|Event|ID)[:\s#]*(\d{4,})', text, re.IGNORECASE)
    if match:
        return match.group(1)
    return None


# ── PDF extraction (pdfplumber) ────────────────────────────────────────────────

def _extract_pdf(content: bytes, filename: str) -> dict:
    try:
        import pdfplumber
    except ImportError:
        return _extract_generic(filename)

    result: dict = {}
    extra: dict = {}
    all_text = ""
    line_items = []

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page_num, page in enumerate(pdf.pages):
            page_text = page.extract_text() or ""
            all_text += page_text + "\n"

            tables = page.extract_tables() or []
            for table in tables:
                if not table:
                    continue

                rows = [[str(c).strip() if c else "" for c in row] for row in table]
                non_empty_rows = [r for r in rows if any(c for c in r)]
                if not non_empty_rows:
                    continue

                # ── Detect if this table is a key-value summary or a data table ──
                first_row = non_empty_rows[0]
                non_empty_cols = sum(1 for c in first_row if c)

                # 2-column tables are usually key-value pairs
                if non_empty_cols == 2 or (len(first_row) >= 2 and not _match_col(first_row[0]) and not _match_col(first_row[1])):
                    for row in non_empty_rows:
                        if len(row) >= 2 and row[0] and row[1]:
                            field = _match_pdf_field(row[0])
                            val = row[1]
                            if field and val:
                                if field not in result:
                                    result[field] = val
                            elif row[0] and row[1] and row[0].lower() not in ("name", "value", "alternative", "item"):
                                extra.setdefault(row[0].title(), val)
                    continue

                # ── Data table with headers → line items ──
                # Detect header row (first row with recognisable column names)
                header_row_idx = None
                col_map = {}
                for i, row in enumerate(non_empty_rows[:5]):
                    mapped = {_match_col(c): j for j, c in enumerate(row) if c and _match_col(c)}
                    if len(mapped) >= 2:
                        col_map = mapped
                        header_row_idx = i
                        break

                if header_row_idx is not None:
                    for row in non_empty_rows[header_row_idx + 1:]:
                        if all(not c for c in row):
                            continue
                        item: dict = {"item_no": str(len(line_items) + 1)}
                        for field, idx in col_map.items():
                            if idx < len(row) and row[idx]:
                                v = row[idx]
                                item[field] = _safe_float(v) if field in ("quantity", "unit_price", "total_price") else v
                        if item.get("quantity") and item.get("unit_price") and not item.get("total_price"):
                            item["total_price"] = round(item["quantity"] * item["unit_price"], 2)
                        if any(v for k, v in item.items() if k != "item_no"):
                            line_items.append(item)
                else:
                    # Treat rows as requirements/content items
                    for row in non_empty_rows:
                        if row[0] and not any(kw in row[0].lower() for kw in ("introduction", "overview", "content", "timing", "currency rule", "section")):
                            val = " | ".join(c for c in row[1:] if c) if len(row) > 1 else ""
                            line_items.append({
                                "item_no": str(len(line_items) + 1),
                                "description": row[0],
                                "notes": val or None,
                            })

    # ── Post-process ──────────────────────────────────────────────────────────
    rfq = _extract_rfq_number(all_text, filename)
    if rfq:
        result.setdefault("rfq_number", rfq)

    case_id = _extract_case_id(filename, all_text)
    if case_id:
        result.setdefault("rfq_case_id", case_id)

    total = sum(i.get("total_price") or 0 for i in line_items if i.get("total_price"))

    return {
        "rfq_number":  result.get("rfq_number"),
        "rfq_case_id": result.get("rfq_case_id"),
        "account":     result.get("owner"),          # company name
        "owner":       result.get("owner"),           # person name (may differ)
        "supplier":    result.get("supplier"),
        "address":     result.get("address"),
        "event_type":  result.get("event_type"),
        "commodity":   result.get("commodity"),
        "issue_date":  result.get("issue_date"),
        "due_date":    result.get("due_date"),
        "currency":    result.get("currency_name", "SAR"),
        "response_id": result.get("response_id"),
        "po_id":       result.get("po_id"),
        "po_number":   result.get("po_number"),
        "agreement_no":result.get("agreement_no"),
        "contact":     result.get("contact"),
        "phone":       result.get("phone"),
        "email":       result.get("email"),
        "total_amount": total or None,
        "line_items":   line_items,
    }


# ── Excel extraction ──────────────────────────────────────────────────────────

def _extract_excel(content: bytes, filename: str) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    summary: dict = {}
    col_indices: dict = {}
    table_start = None
    extra: dict = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
        non_empty = [c for c in row if c is not None]
        if not non_empty:
            continue

        # Table header detection
        if table_start is None:
            hdrs = [str(c).strip() if c else "" for c in row]
            mapped = {_match_col(h): i for i, h in enumerate(hdrs) if h and _match_col(h)}
            if len(mapped) >= 2:
                col_indices = mapped
                table_start = row_idx + 1
                continue

        # Key-value scanning
        for i, cell in enumerate(row):
            if not cell:
                continue
            field = _match_pdf_field(str(cell))
            if field and i + 1 < len(row) and row[i + 1] is not None:
                summary.setdefault(field, str(row[i + 1]).strip())

    line_items = []
    if table_start:
        for row in ws.iter_rows(min_row=table_start, values_only=True):
            if all(c is None for c in row):
                break
            item: dict = {"item_no": str(len(line_items) + 1)}
            for field, idx in col_indices.items():
                if idx < len(row) and row[idx] is not None:
                    v = row[idx]
                    item[field] = _safe_float(v) if field in ("quantity", "unit_price", "total_price") else str(v).strip()
            if item.get("quantity") and item.get("unit_price") and not item.get("total_price"):
                item["total_price"] = round(item["quantity"] * item["unit_price"], 2)
            if any(v for k, v in item.items() if k != "item_no"):
                line_items.append(item)

    rfq = _extract_rfq_number("", filename)
    total = sum(i.get("total_price") or 0 for i in line_items if i.get("total_price"))

    return {
        "rfq_number":  summary.get("rfq_number") or rfq,
        "rfq_case_id": summary.get("rfq_case_id"),
        "account":     summary.get("owner") or summary.get("account"),
        "owner":       summary.get("owner"),
        "address":     summary.get("address"),
        "event_type":  summary.get("event_type"),
        "commodity":   summary.get("commodity"),
        "issue_date":  summary.get("issue_date"),
        "due_date":    summary.get("due_date"),
        "currency":    summary.get("currency_name", "SAR"),
        "response_id": summary.get("response_id"),
        "po_id":       summary.get("po_id"),
        "po_number":   summary.get("po_number"),
        "agreement_no":summary.get("agreement_no"),
        "line_items":  line_items,
        "total_amount": total or None,
    }


# ── CSV extraction ─────────────────────────────────────────────────────────────

def _extract_csv(content: bytes, filename: str) -> dict:
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"line_items": [], "extra_fields": {}}

    headers = [h.strip() for h in rows[0]]
    col_indices = {_match_col(h): i for i, h in enumerate(headers) if h and _match_col(h)}
    line_items = []

    for row_idx, row in enumerate(rows[1:], start=1):
        item: dict = {"item_no": str(row_idx)}
        for field, idx in col_indices.items():
            if idx < len(row) and row[idx].strip():
                v = row[idx].strip()
                item[field] = _safe_float(v) if field in ("quantity", "unit_price", "total_price") else v
        if not item.get("description") and row:
            item["description"] = row[0].strip()
        if item.get("quantity") and item.get("unit_price") and not item.get("total_price"):
            item["total_price"] = round(item["quantity"] * item["unit_price"], 2)
        if any(v for k, v in item.items() if k != "item_no"):
            line_items.append(item)

    total = sum(i.get("total_price") or 0 for i in line_items if i.get("total_price"))
    return {
        "rfq_number":  _extract_rfq_number("", filename),
        "line_items":  line_items,
        "total_amount": total or None,
    }


# ── Generic fallback ──────────────────────────────────────────────────────────

def _extract_generic(filename: str) -> dict:
    return {
        "rfq_number":  _extract_rfq_number("", filename),
        "line_items":  [],
        "extra_fields": {},
    }


# ── Public API ────────────────────────────────────────────────────────────────

def extract(content: bytes, suffix: str, filename: str) -> dict:
    """
    Extract structured fields from an uploaded document.
    Returns dict with: rfq_number, account, supplier, issue_date, due_date,
                       currency, total_amount, line_items, extra_fields (dict)
    """
    try:
        if suffix in (".pdf",):
            result = _extract_pdf(content, filename)
        elif suffix in (".xlsx", ".xls"):
            result = _extract_excel(content, filename)
        elif suffix == ".csv":
            result = _extract_csv(content, filename)
        else:
            result = _extract_generic(filename)
    except Exception:
        result = _extract_generic(filename)

    if not result.get("currency"):
        result["currency"] = "SAR"
    if not result.get("extra_fields"):
        result["extra_fields"] = {}

    return result
